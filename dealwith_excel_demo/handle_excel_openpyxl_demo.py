# coding:utf-8  fxb_qzyx@163.com
"""
   使用第三方库(openpyxl)：来处理excel
      注意事项：1.能够被操作的excel格式为：xlsx、xlsm
               2.在向多行单元格写入数据时，单元格的原样式被修改了
"""
import os
import traceback

from openpyxl import load_workbook


class ParseExcel(object):
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path

        if not os.path.exists(self.excel_file_path):
            raise FileNotFoundError("[error] file not exist!", self.excel_file_path)
        try:
            # 创建excel工作簿对象
            self.wb_obj = load_workbook(excel_file_path)

        except Exception as e:
            traceback.print_exc()

    def getCell(self, sheet_name, row, col):
        """
        :brief 获取指定的单元格值
        :param sheet_name: excel的sheet页名称
        :param row: 单元格行号
        :param col: 单元格列号
        :return: 返回获取的单元格值
        """
        sheet_obj = self.wb_obj.get_sheet_by_name(sheet_name)
        cell_data = sheet_obj.cell(row=row, column=col).value
        return cell_data

    def setCell(self, sheet_name, row, col, val):
        """
        :breif 给指定单元格写入数据
        :param sheet_name: sheet页名称
        :param row: 单元格行号
        :param col: 单元格列号
        :param val: 被写入的数据
        :return: None
        """

        sheet_obj = self.wb_obj.get_sheet_by_name(sheet_name)
        sheet_obj.cell(row=row, column=col, value=val)

    def getRange(self, sheet_name, min_row, min_col, max_row, max_col):
        """
        :brief 获取多行单元格的数据(包括一行)，以列表的形式返回
        :param sheet_name: sheet页名称
        :param min_row: 左上角的行号
        :param min_col: 左上角的列号
        :param max_row: 右下角的行号
        :param max_col: 右下角的列号
        :return: 以列表的形式返回多行单元格的数据
        """
        sheet_obj = self.wb_obj.get_sheet_by_name(sheet_name)
        iter_rows = sheet_obj.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col)
        data_list = []
        for row_obj in iter_rows:
            row_data_list = []
            for row_cell in row_obj:
                row_data_list.append(row_cell.value)
            data_list.append(row_data_list)

        if len(data_list) == 1:  # 若只有一行数据，则降成一维列表
            data_list = data_list[0]

        return data_list

    def setRange(self, sheet_name, min_row, min_col, max_row, max_col, data_list):
        """
        :brief 将数据写入多行单元格(包括一行)
        :param sheet_name: sheet页名称
        :param min_row: 左上角的行号
        :param min_col: 左上角的列号
        :param max_row: 右下角的行号
        :param max_col: 右下角的列号
        :param data_list: 要写入的数据(列表类型)
        :return: None
        """
        sheet_obj = self.wb_obj.get_sheet_by_name(sheet_name)
        # iter_rows:始终是二维
        iter_rows = sheet_obj.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col)
        # 降维：降成一维
        iter_rows = [cell_obj for row_obj in iter_rows for cell_obj in row_obj]
        data_list = [cell_data for row_data_list in data_list for cell_data in row_data_list]
        # 将数据写入对应的单元格
        for cell_obj, cell_data in zip(iter_rows, data_list):
            cell_obj.value = cell_data

    def getSheetNames(self):
        """
        :brief 获取excel的所有sheet页名称
        :return: 以列表的形式返回所有的sheet页
        """
        sheet_name_list = self.wb_obj.get_sheet_names()
        return sheet_name_list

    def getMaxRowNum(self, sheet_name):
        """
        :brief 获取指定sheet页存在数据的最大行号
        :param sheet_name: sheet页名称
        :return: 返回最大行号
        """
        sheet_obj = self.wb_obj.get_sheet_by_name(sheet_name)
        max_row = sheet_obj.max_row
        return max_row

    def getMaxColumnNum(self, sheet_name):
        """
          :brief 获取指定sheet页存在数据的最大列号
          :param sheet_name: sheet页名称
          :return: 返回最大列号
          """
        sheet_obj = self.wb_obj.get_sheet_by_name(sheet_name)
        max_column = sheet_obj.max_column
        return max_column

    def save(self):
        """
        :brief 保存excel文件
        :return: None
        """
        self.wb_obj.save(self.excel_file_path)


if __name__ == "__main__":
    excel_file_path = "./excel_file/ExcelData.xlsx"
    parse_excel = ParseExcel(excel_file_path)
    # print(parse_excel.getCell('massage', 2, 1))
    # print(parse_excel.getSheetNames())
    # parse_excel.setCell('massage', 2, 1, "大神")
    # print('maxrow >>>', parse_excel.getMaxRowNum('massage'))
    # print('maxcolumn >>>', parse_excel.getMaxColumnNum('massage'))
    # print(parse_excel.getRange('massage', 1, 1, 3, 3))
    parse_excel.setRange('massage', 10, 1, 11, 3, [['a', 'b', 'c'], ['d', 'e', 'f']])
    parse_excel.save()
